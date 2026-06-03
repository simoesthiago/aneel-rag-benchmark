"""
office.py — Extração de texto plano de documentos Office (DOCX, XLSX).

Usado por manuais da ANEEL que vêm nesses formatos. Estratégias baseline
(formato_saida="texto"): python-docx (parágrafos) e openpyxl (células).
As estratégias Markdown são `mammoth_md.py` (DOCX) e `xlsx_markdown.py` (XLSX).
"""

import io


def extrair_texto_docx(conteudo: bytes) -> dict:
    """Extrai texto de um arquivo DOCX (parágrafos concatenados)."""
    try:
        from docx import Document
    except ImportError:
        raise RuntimeError(
            "python-docx não está instalado. Instale com: pip install python-docx"
        )

    doc = Document(io.BytesIO(conteudo))
    partes = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    texto_final = "\n\n".join(partes)
    qualidade = 1.0 if len(texto_final) > 1000 else 0.5

    return {
        "texto": texto_final,
        "num_paginas": None,
        "qualidade_extracao": qualidade,
        "formato_saida": "texto",
        "extrator": "python_docx",
    }


def extrair_texto_xlsx(conteudo: bytes) -> dict:
    """Extrai texto de planilha XLSX (células não vazias, por aba)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            "openpyxl não está instalado. Instale com: pip install openpyxl"
        )

    wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    partes: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None and str(cell).strip():
                    partes.append(str(cell).strip())
    wb.close()

    texto_final = "\n".join(partes)
    qualidade = 1.0 if len(texto_final) > 1000 else 0.5

    return {
        "texto": texto_final,
        "num_paginas": None,
        "qualidade_extracao": qualidade,
        "formato_saida": "texto",
        "extrator": "openpyxl",
    }
